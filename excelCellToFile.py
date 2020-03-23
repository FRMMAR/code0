# -*- coding: utf-8 -*-
"""
Created on Mon Mar 23 17:13:58 2020

@author: m

put the cell of excel into another file

"""

import os
import openpyxl
import re

def read_excel_write_file(workingdir, filename, htmlname):
    os.chdir(workingdir)
    wb = openpyxl.load_workbook(filename)
    ws = wb['Sheet1']
    
    with open(htmlname, 'rb') as f:
        f_content = f.read().decode()
        
    for a in range(2,153):
        cellvalue_num = ws.cell(row=a, column=1).value
        cellvalue_name = ws.cell(row=a, column=6).value
        name = str(cellvalue_num) + cellvalue_name +'.html'
        cellvalue_link = ws.cell(row=a, column=8).value
        cellvalue_long = str(ws.cell(row=a, column=3).value)
        cellvalue_lati = str(ws.cell(row=a, column=4).value)
        print(name, cellvalue_link,cellvalue_long, cellvalue_lati)
        try:
            new_con = f_content.replace('zhelishidizhi',cellvalue_link).replace('zhelishijingdu',cellvalue_long).replace('zhelishiweidu',cellvalue_lati)
            with open(name,'a') as newhtml:
                newhtml.write(new_con)
        except Exception as e:
            print(e)
            pass
        continue

def main():
    read_excel_write_file(workingdir, filename, htmlname)
    
if __name__ == '__main__':
    workingdir = r'\home\m\a'
    filename = r'1.xlsx'
    htmlname = r'1.html'
    main()
